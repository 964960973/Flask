import json
import time
from openpyxl import load_workbook
import threading
import jsonpath
import requests
import redis
from  lxml import etree
from openpyxl.drawing.image import Image
import os
from PIL import Image as image_change
from selenium.webdriver.common.by import By
from selenium import webdriver
import redis
redis_client = redis.Redis(host='127.0.0.1', port=6379)

def get_url_link():
    chrome_options = webdriver.ChromeOptions()
    chrome_driver = r"./chromedriver.exe"
    driver = webdriver.Chrome(options=chrome_options, executable_path=chrome_driver)
    name = input('输入你需要获取的职位')
    for page in range(1,11):#打断点 人工过验证码
        driver.get(f'https://www.zhipin.com/web/geek/job?query={name}&page={page}')  # 打开网址.
        time.sleep(5)
        try:
            url_list = driver.find_elements(By.XPATH,'//div[@class="job-card-body clearfix"]/a')
            diqu_list = driver.find_elements(By.XPATH,'//span[@class="job-area"]')
            zhi_wei_list = driver.find_elements(By.XPATH,'//span[@class="job-name"]')
            money_list = driver.find_elements(By.XPATH,'//div[@class="job-info clearfix"]')
            jishu_list = driver.find_elements(By.XPATH,'//div[@class="job-card-footer clearfix"]/ul')
            daiyu_list = driver.find_elements(By.XPATH,'//div[@class="info-desc"]')
            q_ye_list = driver.find_elements(By.XPATH,'//ul[@class="company-tag-list"]/li')
        except:
            print('匹配出现问题')
            continue
        data = {}
        try:
            for url,diqu,zhiwei,jishu,daiyu,q_ye,money in zip(url_list,diqu_list,zhi_wei_list,jishu_list,daiyu_list,q_ye_list,money_list):
                get_url = str(url.get_attribute('href')).split('html?')[0] + 'html'
                a = diqu.text
                b = zhiwei.text
                c = jishu.text
                d = daiyu.text
                e = q_ye.text
                f = money.text
                data['详情链接'] = get_url
                data['所在地区'] = a
                data['职位'] = b
                data['所需技术'] = c
                data['福利待遇'] = d
                data['企业信息'] = e
                data['工资、其他信息'] = f
                redis_client.hset('boss_pa_g',get_url,json.dumps(data))
                print(f'存储成功data=={data}')
        except:
            print('循环出现问题')
            continue
# get_url_link()
def get_image():
    while True:
        try:
            image_url = redis_client.lpop("yoto_user_image_list").decode("utf-8")
            num = redis_client.llen("yoto_user_image_list")
            image_name = image_url.split("rs=")[1]
            headers = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36"
            }
            response = requests.get(image_url, headers=headers,proxies="").content
            with open(f"./image/{image_name}.jpg", "wb") as f:
                f.write(response)
                print(f"图片保存成功，还剩余{num}张")
        except:
            print(f"图片保存失败 image_url = {image_url}")
            redis_client.rpush("yoto_user_image_list",image_url)
            continue


def data_excel():
  wb = load_workbook(f'./测试文件.xlsx') # 加载一个空excel
  sheet = wb.active
  url = redis_client.hkeys(f'boss_pa_g')
  sheet.cell(row=1, column=1).value = "详情链接"
  sheet.cell(row=1, column=2).value = "所在地区"
  sheet.cell(row=1, column=3).value = "职位"
  sheet.cell(row=1, column=4).value = "所需技术"
  sheet.cell(row=1, column=5).value = "福利待遇"
  sheet.cell(row=1, column=6).value = "企业信息"
  sheet.cell(row=1, column=7).value = "工资、其他信息"
  i = 2
  for key in url:
      try:
          data_info = redis_client.hget('boss_pa_g', key)
          info_str = data_info.decode("utf-8")
          info = json.loads(info_str)
          sheet.cell(row=i, column=1).value = info["详情链接"]
          sheet.cell(row=i, column=2).value = info["所在地区"]
          sheet.cell(row=i, column=3).value = info["职位"]
          sheet.cell(row=i, column=4).value = info["所需技术"]
          sheet.cell(row=i, column=5).value = info["福利待遇"]
          sheet.cell(row=i, column=6).value = info["企业信息"]
          sheet.cell(row=i, column=7).value = info["工资、其他信息"]
          i += 1
          print(f"数据存储成功 url == {key}")
          wb.save(f'广州_爬虫.xlsx')

      except:
          break
data_excel()
def addImgToExcel(excelPath, sheetName, imgWidth, imgHight, columnWidth,rowHight):
    # 加载Excel文件并读取指定Sheet
    wb = load_workbook(excelPath)
    ws = wb[sheetName]
    # 获取整个sheet的最大行数
    Max_RowNumber = ws.max_row
    # 从表格第二行开始插入图片
    for rowid in range(2,Max_RowNumber+1):
        try:
            # 设置行高(除去第一行)
            ws.row_dimensions[rowid].height = rowHight
            # 取到第N行第N列的值（图片的全路径）
            # picPath = "./image1/B00EVB04NC.jpg"
            picPathColunmNumber = 5  # 图片地址在第几列
            writeColunm = 71   # 图片需要插在第几列  要用对应字母的数字形式
            name = ""
            for i in range(1,12):
                picPath = ws.cell(rowid, picPathColunmNumber).value
                img_change = image_change.open(picPath).convert("RGB")
                img_change.save(picPath, 'jpeg')
                # 对得到的路径进行判断
                if picPath and os.path.exists(picPath):
                    # 设置图片尺寸大小
                    img = Image(picPath)
                    img.width = imgWidth
                    img.height = imgHight
                    # 写入图片
                    if writeColunm > 90:  # excel插入的名称对应的数字大于90的相当于名称前面多个A
                        writeColunm = 66
                        name = "A"
                    writeColunmName = name + chr(writeColunm)
                    ws.add_image(img, str(writeColunmName) + str(rowid))
                    # 列宽
                    ws.column_dimensions[writeColunmName].width = columnWidth
                    print(f"插入图片成功，i= {rowid},剩余{10000-rowid}个")
                    picPathColunmNumber += 3
                    writeColunm += 3
                else:
                    print("没有图片")
                    continue

        except:
            continue
    wb.save(excelPath)
    return


def run():
    get_image()

def main():
  print(f'主线程开始时间：{time.strftime("%Y-%m-%d %H:%M:%S")}')
  # 初始化3个线程，传递不同的参数
  # for i in range(4):
  t1 = threading.Thread(target=run)
  t2 = threading.Thread(target=run)
  t3 = threading.Thread(target=run)
  # 开启三个线程
  t1.start()
  t2.start()
  t3.start()
  # 等待运行结束
  t1.join()
  t2.join()
  t3.join()

if __name__ == '__main__':
    pass
    # get_url_link()#获取主页面链接
    # get_url() #获取详情页信息
    # data_excel()#插入ex表格
    # main()#线程下载图片
    # addImgToExcel(f'./yuto_user_info.xlsx', "Sheet1", 65, 71, 8, 55)#插入图片
