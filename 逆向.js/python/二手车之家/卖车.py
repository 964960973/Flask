import json
import redis
import requests
import threading

from lxml import etree
from lxml.doctestcompare import strip

redis_client = redis.Redis(host='127.0.0.1', port=6379)


def get_ip():
  # 提取代理API接口，获取1个代理IP
  api_url = "http://v2.api.juliangip.com/dynamic/getips?num=1&pt=1&result_type=text&split=1&trade_no=1324587523689164&sign=353a9894a22c17ae524f0ae6e412b44c"

  # 获取API接口返回的代理IP
  proxy_ip = requests.get(api_url).text

  # 用户名密码认证(动态代理/独享代理)
  username = "18038573677"
  password = "jti9iyhc"
  proxies = {
    "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip},
    "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip}
  }
  return proxies


def image():
    while True:
        try:
            image_url = redis_client.lpop('e_s_c_imamge').decode("utf-8")
            ids = str(image_url).split('.jpg__')[1]
            url = str(image_url).split('.jpg__')[0] + '.jpg'
            headers = {
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36"
            }
            response = requests.get(url, headers=headers).content
            with open(f"./image/{ids}.jpg", "wb") as f:
                f.write(response)
                print(f'图片保存成功image==={url}')
        except:
            print('图片已全部保存完毕')
            break


def get_link():
    for page in range(10,22):
        url = f"https://www.che168.com/china/guangqichuanqi/a0_0msdgscncgpi1ltocsp{page}exx0/?pvareaid=102179#currengpostion"
        payload = {}
        headers = {
              'authority': 'www.che168.com',
              'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
              'accept-language': 'zh-CN,zh;q=0.9',
              'cache-control': 'no-cache',
              'cookie': 'fvlid=1675930509561UAoWbJ7fzKc7; sessionid=58058755-dd7a-489d-8730-44dbfbd2d7cb; sessionip=116.21.238.193; area=440111; sessionvisit=cea9cada-aa45-49b7-8248-869e6322cfee; sessionvisitInfo=58058755-dd7a-489d-8730-44dbfbd2d7cb|www.baidu.com|0; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1675930510; che_sessionid=160CD66E-5E81-40D5-9A60-EB03DA50C669%7C%7C2023-02-09+16%3A15%3A09.521%7C%7Cwww.baidu.com; che_sessionvid=88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8; userarea=440100; qimo_seosource_7a783820-ec84-11ec-b95f-79694d4df285=%E7%AB%99%E5%86%85; qimo_seokeywords_7a783820-ec84-11ec-b95f-79694d4df285=; qimo_xstKeywords_7a783820-ec84-11ec-b95f-79694d4df285=; href=http%3A%2F%2Fwww.che168.com%2Fguangzhou%2F; accessId=7a783820-ec84-11ec-b95f-79694d4df285; pageViewNum=1; listuserarea=440100; UsedCarBrowseHistory=0%3A46520568; carDownPrice=1; ahpvno=7; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1675930591; ahuuid=B3DDB9B2-8EE3-4151-84DA-E4CF76CA2F55; showNum=7; v_no=7; visit_info_ad=160CD66E-5E81-40D5-9A60-EB03DA50C669||88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8||-1||-1||7; che_ref=www.baidu.com%7C0%7C0%7C0%7C2023-02-09+16%3A16%3A30.720%7C2023-02-09+16%3A15%3A09.521; sessionuid=58058755-dd7a-489d-8730-44dbfbd2d7cb',
              'pragma': 'no-cache',
              'referer': 'https://www.che168.com/guangzhou/list/',
              'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
              'sec-ch-ua-mobile': '?0',
              'sec-ch-ua-platform': '"Windows"',
              'sec-fetch-dest': 'document',
              'sec-fetch-mode': 'navigate',
              'sec-fetch-site': 'same-origin',
              'sec-fetch-user': '?1',
              'upgrade-insecure-requests': '1',
              'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
        }
        response = requests.request("GET", url, headers=headers, data=payload).text
        tree = etree.HTML(response)
        goods_url = tree.xpath('//li[@class="cards-li list-photo-li "]/a/@href')
        for i in goods_url:
            if 'www.che' not in i:
                get_url = 'https://www.che168.com' + i
            else:
                get_url = 'https:' + i
            redis_client.lpush('e_s_c',get_url)
            print(f'存储成功url==={get_url}')


def liang_dian(ids,proxies):
    url = f"https://apipcmusc.che168.com/v1/car/getusedcaroptiondata?callback=getUsedCarOptionDataCallback&_appid=2sc.m&infoid={ids}"
    payload = {}
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Cookie': 'fvlid=1675930509561UAoWbJ7fzKc7; sessionid=58058755-dd7a-489d-8730-44dbfbd2d7cb; sessionip=116.21.238.193; area=440111; sessionvisit=cea9cada-aa45-49b7-8248-869e6322cfee; sessionvisitInfo=58058755-dd7a-489d-8730-44dbfbd2d7cb|www.baidu.com|0; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1675930510; che_sessionid=160CD66E-5E81-40D5-9A60-EB03DA50C669%7C%7C2023-02-09+16%3A15%3A09.521%7C%7Cwww.baidu.com; che_sessionvid=88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8; userarea=440100; listuserarea=440100; showNum=18; sessionuid=58058755-dd7a-489d-8730-44dbfbd2d7cb; v_no=18; visit_info_ad=160CD66E-5E81-40D5-9A60-EB03DA50C669||88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8||-1||-1||18; che_ref=www.baidu.com%7C0%7C0%7C0%7C2023-02-09+16%3A41%3A25.027%7C2023-02-09+16%3A15%3A09.521; ahpvno=19; ahuuid=34A22536-1A67-4877-8CD0-DCFCC7D0014B; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1675932099',
        'Pragma': 'no-cache',
        'Referer': 'https://www.che168.com/',
        'Sec-Fetch-Dest': 'script',
        'Sec-Fetch-Mode': 'no-cors',
        'Sec-Fetch-Site': 'same-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"'
    }
    response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
    json_data = '{' + str(response).split('({')[1].split('})')[0] + '}'
    json_list = json.loads(json_data)
    l_dian_list = []
    oss = json_list['result']
    for i in oss:
        da = i['optionname']
        l_dian_list.append(da)
    return l_dian_list




def good_link():
    proxies = get_ip()
    while True:
        # url = 'https://www.che168.com/dealer/67783/46357417.html?pvareaid=100519&userpid=440000&usercid=440100&offertype=&offertag=0&activitycartype=0'
        url = redis_client.lpop('e_s_c').decode("utf-8")
        payload = {}
        headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Cookie': 'userarea=0; listuserarea=0; ahpvno=1; fvlid=1676080299405YKEDMCwyv0U1; ahuuid=42FECA2C-6873-41D8-B472-7D1D96DC7F68; sessionid=ae34c623-79c8-4d0d-9f0f-b777ae9cef93; sessionip=113.68.39.206; area=440111; sessionvisit=727e041d-dce0-452f-996b-f5344d3bbc63; sessionvisitInfo=ae34c623-79c8-4d0d-9f0f-b777ae9cef93||102179; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1676080300; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1676080300; showNum=1; che_sessionid=90088CE6-B4E2-4C84-953B-70598C72FA11%7C%7C2023-02-11+09%3A51%3A39.819%7C%7C0; v_no=1; visit_info_ad=90088CE6-B4E2-4C84-953B-70598C72FA11||E3DE1B5D-BE76-455B-9B37-A129250DA690||-1||-1||1; che_ref=0%7C0%7C0%7C0%7C2023-02-11+09%3A51%3A39.819%7C2023-02-11+09%3A51%3A39.819; che_sessionvid=E3DE1B5D-BE76-455B-9B37-A129250DA690; sessionuid=ae34c623-79c8-4d0d-9f0f-b777ae9cef93',
            'Pragma': 'no-cache',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"'
        }
        try:
            response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
            if '您跑的太快了' in response:
                print(f'更换代理访问过快==={url},proxies==={proxies}')
                redis_client.rpush('e_s_c', url)
                time.sleep(5)
                proxies = get_ip()
                continue
            ids = str(url.split('.html?')[0].split('/')[-1])
            tree = etree.HTML(response)
            data = {}
            d_a_list = []
            name = tree.xpath('//h3[@class="car-brand-name"]/text()')[0]#车名
            ji_keys = tree.xpath('//ul[@class="brand-unit-item fn-clear"]/li/p/text()')#车辆基本信息keys
            ji_values = tree.xpath('//ul[@class="brand-unit-item fn-clear"]/li/h4/text()')#车辆基本信息values
            jia_ge = tree.xpath('//span[@class="price-present"]/text()')[0]#价格
            d_a_keys = tree.xpath('//ul[@class="basic-item-ul"]/li/span/text()') #车辆档案基本信息keys
            d_a_values = tree.xpath('//ul[@class="basic-item-ul"]/li/text()')#车辆档案基本信息values
            get_image = 'https:' + tree.xpath('//div[@class="cards-heng-img"]/img/@src')[0] + '__' + ids#车辆图片网址
            liangdian = liang_dian(ids,proxies) #车辆配置亮点
            redis_client.lpush('e_s_c_imamge',get_image)
            data['image'] = ids
            print(f'图片存储成功image_url==={get_image}')
            for i in d_a_keys:
                if i == '维修保养':
                    continue
                elif i == '出险查询':
                    continue
                else:
                    d_a_list.append(i)
            for o,p in zip(d_a_list,d_a_values):
                data[o] = p

            for k,j in zip(ji_keys,ji_values):
                data[k] = j
            data['车名'] = name
            data['价格'] = jia_ge
            data['亮点'] = liangdian
            redis_client.hset(f"cut_hash", ids, json.dumps(data))
            print(f'存储成功data==={data}')
        except:
            print(f'出现异常url==={url}')
            time.sleep(3)
            proxies = get_ip()
            # redis_client.rpush('e_s_c',url)
            continue
# good_link()
# get_link()
def get_data_ex():
    wb = load_workbook('./测试文件.xlsx')
    zhuye_url = redis_client.hkeys(f'cut_hash')
    row_num = 1
    sheet = wb['Sheet1']
    for key in zhuye_url:
        i = 2
        # try:
        data_info = redis_client.hget('cut_hash',key)
        info_str = data_info.decode("utf-8")
        info = json.loads(info_str)
        for aa in info.keys():
            data = strip(aa)
            oo = strip(str(info[aa]))
            sheet.cell(row=row_num, column=i).value = str(data)
            sheet.cell(row=row_num+1, column=i).value = str(oo)
            i += 1
        row_num += 2
        #     p_l = info['revirew']
        #     column = 6
        #     opp = 7
        #     for image_info in p_l:
        #         sheet.cell(row=i,column=column).value = str(image_info.keys()).split('])')[0].split('([')[1]
        #         sheet.cell(row=i,column=opp).value = str(image_info.values()).split('])')[0].split('([')[1]
        #         column += 2
        #         opp += 2
        #     biao_qian_list = 12
        #     if len(info['biao_qian']) == 0:
        #         sheet.cell(row=i, column=biao_qian_list).value = '无标签'
        #     for biao_qian in info['biao_qian']:
        #         sheet.cell(row=i, column=biao_qian_list).value = biao_qian
        #         biao_qian_list += 1
        #     image_a = 20
        #     image_url = info['image']
        #     image_name = image_url.split("x-signature=")[1]
        #     sheet.cell(row=i, column=image_a).value = f"./image/{image_name}.jpg"
        # except:
        #     print('异常')
        # i += 1
        print(f"数据存储成功 url == {key}")
    wb.save(f'./先驱者.xlsx')

# get_data_ex()




from openpyxl import load_workbook
import os
from openpyxl.drawing.image import Image
import time

times = time.localtime().tm_year
ti = time.localtime().tm_mon
t = time.localtime().tm_mday
year = str(times)+'_'+str(ti)+'_'+str(t)

def addImgToExcel(excelPath, sheetName, picPathColunmNumber, writeColunmName, imgWidth, imgHight, columnWidth,
                  rowHight):
    # 加载Excel文件并读取指定Sheet
    wb = load_workbook(excelPath)
    ws = wb[sheetName]
    # 获取整个sheet的最大行数
    Max_RowNumber = ws.max_row
    # 从表格第二行开始插入图片
    for rowid in range(2,Max_RowNumber+1):
        try:
            # 设置行高(除去第一行)
            ws.row_dimensions[rowid].height = rowHight
            # 取到第N行第N列的值（图片的全路径）
            # picPath = "./image1/B00EVB04NC.jpg"
            picPath = ws.cell(rowid, picPathColunmNumber).value
            # 对得到的路径进行判断
            # 对得到的路径进行判断writeColunm
            if picPath and os.path.exists(picPath):
                # 设置图片尺寸大小
                img = Image(picPath)
                img.width = imgWidth
                img.height = imgHight
                # 写入图片
                ws.add_image(img, str(writeColunmName) + str(rowid))
                # 列宽
                ws.column_dimensions[writeColunmName].width = columnWidth
                print(f"插入图片成功，i= {rowid},类目==={sheetName}")
            else:
                print(f"插入图片失败类目==={sheetName}")
        except:
            continue
    wb.save(excelPath)

    return

if __name__ == '__main__':
    addImgToExcel(f'./先驱者.xlsx', 'Sheet1', 2, 'A', 65, 71, 8, 55)


def run():
    # good_link()
    image()

def main():
  print(f'主线程开始时间：{time.strftime("%Y-%m-%d %H:%M:%S")}')
  # 初始化3个线程，传递不同的参数
  # for i in range(4):0
  t1 = threading.Thread(target=run)
  t2 = threading.Thread(target=run)
  t3 = threading.Thread(target=run)
  # t4 = threading.Thread(target=run)

  # 开启三个线程
  t1.start()
  t2.start()
  t3.start()
  # t4.start()
  # 等待运行结束
  t1.join()
  t2.join()
  t3.join()
  # t4.join()
  print(f'主线程结束时间：{time.strftime("%Y-%m-%d %H:%M:%S")}')
# main()
# get_link()
# get_data_ex()