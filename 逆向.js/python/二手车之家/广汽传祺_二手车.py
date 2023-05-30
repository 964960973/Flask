import json
import redis
import requests
import threading
from openpyxl import load_workbook
import os
from openpyxl.drawing.image import Image
import time
from lxml import etree
from lxml.doctestcompare import strip

redis_client = redis.Redis(host='127.0.0.1', port=6379)


def get_ip():
  # 提取代理API接口，获取1个代理IP
  api_url = "http://v2.api.juliangip.com/dynamic/getips?num=1&pt=1&result_type=text&split=1&trade_no=1324587523689164&sign=353a9894a22c17ae524f0ae6e412b44c"

  # 获取API接口返回的代理IP
  proxy_ip = requests.get(api_url).text

  # 用户名密码认证(动态代理/独享代理)
  username = "18038573677"
  password = "jti9iyhc"
  proxies = {
    "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip},
    "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip}
  }
  return proxies


def image():
    while True:
        try:
            image_url = redis_client.lpop('e_s_c_imamge').decode("utf-8")
            ids = str(image_url).split('.jpg')[1]
            url = str(image_url).split('.jpg')[0] + '.jpg'
            headers = {
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36"
            }
            response = requests.get(url, headers=headers).content
            with open(f"./image/{ids}.jpg", "wb") as f:
                f.write(response)
                print(f'图片保存成功image==={url}')
        except:
            print('图片已全部保存完毕')
            break


def get_link():
    for page in range(1,22):
        url = f"https://www.che168.com/china/guangqichuanqi/a0_0msdgscncgpi1ltocsp{page}exx0/?pvareaid=102179#currengpostion"
        payload = {}
        headers = {
              'authority': 'www.che168.com',
              'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
              'accept-language': 'zh-CN,zh;q=0.9',
              'cache-control': 'no-cache',
              'cookie': 'fvlid=1675930509561UAoWbJ7fzKc7; sessionid=58058755-dd7a-489d-8730-44dbfbd2d7cb; sessionip=116.21.238.193; area=440111; sessionvisit=cea9cada-aa45-49b7-8248-869e6322cfee; sessionvisitInfo=58058755-dd7a-489d-8730-44dbfbd2d7cb|www.baidu.com|0; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1675930510; che_sessionid=160CD66E-5E81-40D5-9A60-EB03DA50C669%7C%7C2023-02-09+16%3A15%3A09.521%7C%7Cwww.baidu.com; che_sessionvid=88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8; userarea=440100; qimo_seosource_7a783820-ec84-11ec-b95f-79694d4df285=%E7%AB%99%E5%86%85; qimo_seokeywords_7a783820-ec84-11ec-b95f-79694d4df285=; qimo_xstKeywords_7a783820-ec84-11ec-b95f-79694d4df285=; href=http%3A%2F%2Fwww.che168.com%2Fguangzhou%2F; accessId=7a783820-ec84-11ec-b95f-79694d4df285; pageViewNum=1; listuserarea=440100; UsedCarBrowseHistory=0%3A46520568; carDownPrice=1; ahpvno=7; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1675930591; ahuuid=B3DDB9B2-8EE3-4151-84DA-E4CF76CA2F55; showNum=7; v_no=7; visit_info_ad=160CD66E-5E81-40D5-9A60-EB03DA50C669||88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8||-1||-1||7; che_ref=www.baidu.com%7C0%7C0%7C0%7C2023-02-09+16%3A16%3A30.720%7C2023-02-09+16%3A15%3A09.521; sessionuid=58058755-dd7a-489d-8730-44dbfbd2d7cb',
              'pragma': 'no-cache',
              'referer': 'https://www.che168.com/guangzhou/list/',
              'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
              'sec-ch-ua-mobile': '?0',
              'sec-ch-ua-platform': '"Windows"',
              'sec-fetch-dest': 'document',
              'sec-fetch-mode': 'navigate',
              'sec-fetch-site': 'same-origin',
              'sec-fetch-user': '?1',
              'upgrade-insecure-requests': '1',
              'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36'
        }
        response = requests.request("GET", url, headers=headers, data=payload).text
        tree = etree.HTML(response)
        goods_url = tree.xpath('//li[@class="cards-li list-photo-li "]/a/@href')
        for i in goods_url:
            if 'www.che' not in i:
                get_url = 'https://www.che168.com' + i
            else:
                get_url = 'https:' + i
            redis_client.lpush('e_s_c',get_url)
            print(f'存储成功url==={get_url}')


def liang_dian(ids,proxies):
    url = f"https://apipcmusc.che168.com/v1/car/getusedcaroptiondata?callback=getUsedCarOptionDataCallback&_appid=2sc.m&infoid={ids}"
    payload = {}
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Cookie': 'fvlid=1675930509561UAoWbJ7fzKc7; sessionid=58058755-dd7a-489d-8730-44dbfbd2d7cb; sessionip=116.21.238.193; area=440111; sessionvisit=cea9cada-aa45-49b7-8248-869e6322cfee; sessionvisitInfo=58058755-dd7a-489d-8730-44dbfbd2d7cb|www.baidu.com|0; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1675930510; che_sessionid=160CD66E-5E81-40D5-9A60-EB03DA50C669%7C%7C2023-02-09+16%3A15%3A09.521%7C%7Cwww.baidu.com; che_sessionvid=88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8; userarea=440100; listuserarea=440100; showNum=18; sessionuid=58058755-dd7a-489d-8730-44dbfbd2d7cb; v_no=18; visit_info_ad=160CD66E-5E81-40D5-9A60-EB03DA50C669||88CFA0E2-FC09-4D92-9989-DA3C1EAE26C8||-1||-1||18; che_ref=www.baidu.com%7C0%7C0%7C0%7C2023-02-09+16%3A41%3A25.027%7C2023-02-09+16%3A15%3A09.521; ahpvno=19; ahuuid=34A22536-1A67-4877-8CD0-DCFCC7D0014B; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1675932099',
        'Pragma': 'no-cache',
        'Referer': 'https://www.che168.com/',
        'Sec-Fetch-Dest': 'script',
        'Sec-Fetch-Mode': 'no-cors',
        'Sec-Fetch-Site': 'same-site',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Not_A Brand";v="99", "Google Chrome";v="109", "Chromium";v="109"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"'
    }
    response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
    json_data = '{' + str(response).split('({')[1].split('})')[0] + '}'
    json_list = json.loads(json_data)
    l_dian_list = []
    oss = json_list['result']
    for i in oss:
        da = i['optionname']
        l_dian_list.append(da)
    return l_dian_list




def good_link():
    proxies = get_ip()
    while True:
        try:
            # url = "https://www.che168.com/dealer/148621/46957237.html?pvareaid=110520&userpid=0&usercid=0&offertype=&offertag=0&activitycartype=0"
            url = redis_client.lpop('e_s_c').decode("utf-8")
        except:
            print('已全部完成')
            return
        # url = "https://www.che168.com/dealer/148621/46957237.html?pvareaid=110520&userpid=0&usercid=0&offertype=&offertag=0&activitycartype=0"
        payload = {}
        headers = {
            'authority': 'www.che168.com',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'no-cache',
            'cookie': 'fvlid=1675930509561UAoWbJ7fzKc7; sessionid=58058755-dd7a-489d-8730-44dbfbd2d7cb; area=440111; Hm_lvt_d381ec2f88158113b9b76f14c497ed48=1675930510; che_sessionid=160CD66E-5E81-40D5-9A60-EB03DA50C669%7C%7C2023-02-09+16%3A15%3A09.521%7C%7Cwww.baidu.com; href=http%3A%2F%2Fwww.che168.com%2Fguangzhou%2F; sessionip=113.68.38.233; carDownPrice=1; listuserarea=0; sessionvisit=1116c1ba-ec62-4357-b6e1-41a73d934290; sessionvisitInfo=58058755-dd7a-489d-8730-44dbfbd2d7cb||100519; che_sessionvid=92426714-68D1-4D00-991B-F4B9FD9ECCBB; userarea=410100; ahpvno=22; UsedCarBrowseHistory=0%3A46957237%2C0%3A46519670%2C0%3A46957657%2C0%3A46345538%2C0%3A46357417%2C0%3A46953045%2C0%3A46321638%2C0%3A43239092%2C0%3A46427122%2C0%3A46558480; ahuuid=2414BBE1-7F44-4BC5-8F3C-C4F422A1DADE; Hm_lpvt_d381ec2f88158113b9b76f14c497ed48=1676603432; showNum=46; v_no=50; visit_info_ad=160CD66E-5E81-40D5-9A60-EB03DA50C669||92426714-68D1-4D00-991B-F4B9FD9ECCBB||-1||-1||50; che_ref=www.baidu.com%7C0%7C0%7C0%7C2023-02-17+11%3A10%3A32.346%7C2023-02-09+16%3A15%3A09.521; sessionuid=58058755-dd7a-489d-8730-44dbfbd2d7cb',
            'pragma': 'no-cache',
            'referer': 'https://www.che168.com/china/guangqichuanqi/a0_0msdgscncgpi1ltocsp2exx0/?pvareaid=102179',
            'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36'
        }
        try:
            response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
            if '您跑的太快了' in response:
                print(f'更换代理访问过快==={url},proxies==={proxies}')
                redis_client.rpush('e_s_c', url)
                time.sleep(5)
                proxies = get_ip()
                continue
            elif '您访问的车辆信息不存在！' in response:
                print('不存在跳过改链接')
                continue
            ids = str(url.split('.html?')[0].split('/')[-1])
            tree = etree.HTML(response)
            data = {}
            try:
                try:
                    name = tree.xpath('//h3[@class="car-brand-name"]/text()')[0]#车名
                except:
                    name = ''
                try:
                    jia_ge = tree.xpath('//span[@class="price-present"]/text()')[0]  # 价格
                except:
                    jia_ge = ''
                try:
                    shang_pan_time = tree.xpath('//ul[@class="basic-item-ul"]/li[1]/text()')[0]#上牌时间
                except:
                    shang_pan_time = ''
                try:
                    q = tree.xpath('//ul[@class="basic-item-ul"]/li[1]/text()')[1]#年检到期
                except:
                    q = ''
                try:
                    pp = tree.xpath('//ul[@class="basic-item-ul"]/li[1]/text()')[2]#发动机
                except:
                    pp = ''
                try:
                    aa = tree.xpath('//ul[@class="basic-item-ul"]/li[2]/text()')[0]#表显里程
                except:
                    aa = ''
                try:
                    ss = tree.xpath('//ul[@class="basic-item-ul"]/li[2]/text()')[1]#保险到期
                except:
                    ss = ''
                try:
                    qq = tree.xpath('//ul[@class="basic-item-ul"]/li[2]/text()')[2]#车辆级别
                except:
                    qq = ''
                try:
                    jj = tree.xpath('//ul[@class="basic-item-ul"]/li[3]/text()')[0]#变速箱
                except:
                    jj = ''
                try:
                    mm = tree.xpath('//ul[@class="basic-item-ul"]/li[3]/text()')[1]#质保到期
                except:
                    mm = ''
                try:
                    la = tree.xpath('//ul[@class="basic-item-ul"]/li[3]/text()')[2]#车身颜色
                except:
                    la = ''
                try:
                    p_l = tree.xpath('//ul[@class="basic-item-ul"]/li[5]/text()')[0]#排量
                except:
                    p_l = ''
                try:
                    g_h = tree.xpath('//ul[@class="basic-item-ul"]/li[5]/text()')[1]#过户次数
                except:
                    g_h = ''
                try:
                    q_d = tree.xpath('//ul[@class="basic-item-ul"]/li[5]/text()')[2]#驱动方式
                except:
                    q_d = ''
                try:
                    f_b_time = tree.xpath('//ul[@class="basic-item-ul"]/li[6]/text()')[0]#发布时间
                except:
                    f_b_time = ''
                try:
                    s_z_d = tree.xpath('//ul[@class="basic-item-ul"]/li[6]/text()')[1]#所在地
                except:
                    s_z_d = ''
                liangdian = liang_dian(ids, proxies)  # 车辆配置亮点
                image = 'https:' + tree.xpath('//div[@class="cards-heng-img"]/img/@src')[0] + ids
            except:
                print(f'匹配有误跳过该网页url==={url}')
                continue
            data['车名'] = name
            data['价格'] = jia_ge
            data['上牌时间'] = shang_pan_time
            data['年检到期'] = q
            data['发动机'] = pp
            data['表显里程'] = aa
            data['保险到期'] = ss
            data['车辆级别'] = qq
            data['变速箱'] = jj
            data['质保到期'] = mm
            data['车身颜色'] = la
            data['排量'] = p_l
            data['过户次数'] = g_h
            data['驱动方式'] = q_d
            data['发布时间'] = f_b_time
            data['所在地'] = s_z_d
            data['亮点'] = liangdian
            data['id'] = ids
            data_json = json.dumps(data)
            redis_client.hset(f"cut_hash", ids,data_json)
            redis_client.lpush('e_s_c_imamge',image)
            print(f'存储成功data==={data}')
        except:
            redis_client.rpush('e_s_c', url)
            proxies = get_ip()
            print(f'出现异常url==={url}')
            continue
def get_data_ex():
    wb = load_workbook(f'./测试文件.xlsx')
    ids = redis_client.hkeys(f'cut_hash')
    sheet = wb['Sheet1']
    sheet.cell(row=1, column=1).value = "参考图片"
    sheet.cell(row=1, column=2).value = "车名"
    sheet.cell(row=1, column=3).value = "价格"
    sheet.cell(row=1, column=4).value = "上牌时间"
    sheet.cell(row=1, column=5).value = "年检到期"
    sheet.cell(row=1, column=6).value = "发动机"
    sheet.cell(row=1, column=7).value = "表显里程"
    sheet.cell(row=1, column=8).value = "保险到期"
    sheet.cell(row=1, column=9).value = "车辆级别"
    sheet.cell(row=1, column=10).value = "变速箱"
    sheet.cell(row=1, column=11).value = "质保到期"
    sheet.cell(row=1, column=12).value = "车身颜色"
    sheet.cell(row=1, column=13).value = "排量"
    sheet.cell(row=1, column=14).value = "过户次数"
    sheet.cell(row=1, column=15).value = "驱动方式"
    sheet.cell(row=1, column=16).value = "发布时间"
    sheet.cell(row=1, column=17).value = "所在地"
    sheet.cell(row=1, column=18).value = "亮点"
    sheet.cell(row=1, column=19).value = "id"
    i = 2
    for key in ids:
        try:
            data_info = redis_client.hget('cut_hash',key)
            info_str = data_info.decode("utf-8")
            info = json.loads(info_str)
            id = info['id']
            sheet.cell(row=i, column=2).value = strip(info["车名"])
            sheet.cell(row=i, column=3).value = info["价格"]
            sheet.cell(row=i, column=4).value = info["上牌时间"]
            sheet.cell(row=i, column=5).value = info["年检到期"]
            sheet.cell(row=i, column=6).value = info["发动机"]
            sheet.cell(row=i, column=7).value = info["表显里程"]
            sheet.cell(row=i, column=8).value = info["保险到期"]
            sheet.cell(row=i, column=9).value = info["车辆级别"]
            sheet.cell(row=i, column=10).value = info["变速箱"]
            sheet.cell(row=i, column=11).value = info["质保到期"]
            sheet.cell(row=i, column=12).value = info["车身颜色"]
            sheet.cell(row=i, column=13).value = info["排量"]
            sheet.cell(row=i, column=14).value = info["过户次数"]
            sheet.cell(row=i, column=15).value = info["驱动方式"]
            sheet.cell(row=i, column=16).value = info["发布时间"]
            sheet.cell(row=i, column=17).value = info["所在地"]
            sheet.cell(row=i, column=18).value = str(info["亮点"])
            sheet.cell(row=i, column=19).value = str(f"./image/{str(id)}.jpg")
            i += 1
            print(f"数据存储成功 url == {key}")
        except:
            continue
    wb.save(f'广汽传祺.xlsx')
# get_data_ex()



times = time.localtime().tm_year
ti = time.localtime().tm_mon
t = time.localtime().tm_mday
year = str(times)+'_'+str(ti)+'_'+str(t)

def addImgToExcel(excelPath, sheetName, picPathColunmNumber, writeColunmName, imgWidth, imgHight, columnWidth,
                  rowHight):
    # 加载Excel文件并读取指定Sheet
    wb = load_workbook(excelPath)
    ws = wb[sheetName]
    # 获取整个sheet的最大行数
    Max_RowNumber = ws.max_row
    # 从表格第二行开始插入图片
    for rowid in range(2,Max_RowNumber+1):
        try:
            # 设置行高(除去第一行)
            ws.row_dimensions[rowid].height = rowHight
            # 取到第N行第N列的值（图片的全路径）
            # picPath = "./image1/B00EVB04NC.jpg"
            picPath = ws.cell(rowid, picPathColunmNumber).value
            # 对得到的路径进行判断
            # 对得到的路径进行判断writeColunm
            if picPath and os.path.exists(picPath):
                # 设置图片尺寸大小
                img = Image(picPath)
                img.width = imgWidth
                img.height = imgHight
                # 写入图片
                ws.add_image(img, str(writeColunmName) + str(rowid))
                # 列宽
                ws.column_dimensions[writeColunmName].width = columnWidth
                print(f"插入图片成功，i= {rowid},类目==={sheetName}")
            else:
                print(f"插入图片失败类目==={sheetName}")
        except:
            continue
    wb.save(excelPath)

    return

if __name__ == '__main__':
    # pass
    addImgToExcel(f'广汽传祺.xlsx', 'Sheet1', 19, 'A', 65, 71, 8, 55)


def run():
    # good_link()
    image()

def main():
  print(f'主线程开始时间：{time.strftime("%Y-%m-%d %H:%M:%S")}')
  # 初始化3个线程，传递不同的参数
  # for i in range(4):0
  t1 = threading.Thread(target=run)
  t2 = threading.Thread(target=run)
  t3 = threading.Thread(target=run)
  # t4 = threading.Thread(target=run)

  # 开启三个线程
  t1.start()
  t2.start()
  t3.start()
  # t4.start()
  # 等待运行结束
  t1.join()
  t2.join()
  t3.join()
  # t4.join()
  print(f'主线程结束时间：{time.strftime("%Y-%m-%d %H:%M:%S")}')
# main()
# get_link()
# get_data_ex()