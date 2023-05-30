import os

from PIL import Image as image_change
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


def addImgToExcel():
    # 加载Excel文件并读取指定Sheet
    writeColunmName = "A"
    wb = load_workbook(f'//XZ-RO366PCF59RI/image_all/all_photo/reviews.xlsx')
    # wb = load_workbook(f'//XZ-RO366PCF59RI/image_all/all_photo/2023_4_11/商品图片展示.xlsx')
    ws = wb["Sheet1"]
    # 获取整个sheet的最大行数
    Max_RowNumber = ws.max_row
    # 从表格第二行开始插入图片
    for rowid in range(2,Max_RowNumber+1):
        try:
            # 设置行高(除去第一行)
            ws.row_dimensions[rowid].height = 55
            # 取到第N行第N列的值（图片的全路径）
            # picPath = "./image1/B00EVB04NC.jpg"
            pic= ws.cell(rowid, 7).value
            picPath = "//XZ-RO366PCF59RI/image_all/image/"+pic+".jpg"
            # 对得到的路径进行判断
            # 对得到的路径进行判断writeColunm
            if picPath and os.path.exists(picPath):
                # 设置图片尺寸大小
                img = Image(picPath)
                img.width = 65
                img.height = 71
                # 写入图片
                ws.add_image(img, str(writeColunmName) + str(rowid))
                # 列宽
                ws.column_dimensions[writeColunmName].width = 8
                print(f"插入图片成功，i= {rowid}")
            else:
                print(f"插入图片失败")
        except:
            continue
    wb.save(f'//XZ-RO366PCF59RI/image_all/all_photo/reviews.xlsx')
        # wb.save(f'//XZ-RO366PCF59RI/image_all/all_photo/2023_4_11/商品图片展示.xlsx')
    return

addImgToExcel()