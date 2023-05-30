import json
import requests
import re
import redis

import json
import time
import re
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
import redis
redis_client = redis.Redis(host='127.0.0.1', port=6379)

def get_ip_1():
  # 提取代理API接口，获取1个代理IP
  api_url = "http://v2.api.juliangip.com/dynamic/getips?num=1&pt=1&result_type=text&split=1&trade_no=1324587523689164&sign=353a9894a22c17ae524f0ae6e412b44c"

  # 获取API接口返回的代理IP
  proxy_ip = requests.get(api_url).text

  # 用户名密码认证(动态代理/独享代理)
  username = "18038573677"
  password = "jti9iyhc"
  proxies = {
    "http": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip},
    "https": "http://%(user)s:%(pwd)s@%(proxy)s/" % {"user": username, "pwd": password, "proxy": proxy_ip}
  }
  return proxies

def login():
    for page in range(1,5):
        url = "https://gz.meituan.com/meishi/rating/pn{page}/"
        payload={}
        headers = {
          'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
          'Accept-Language': 'zh-CN,zh;q=0.9',
          'Cache-Control': 'no-cache',
          'Connection': 'keep-alive',
          'Cookie': 'uuid=d3241f9aa4b04b23b78a.1678695846.1.0.0; _lxsdk_cuid=186da11d302c8-0a33418a36f0f-26031951-1fa400-186da11d302c8; WEBDFPID=70w500y31wvv59vu125w6u2vv1ux879w813wz88w30w979583190v6u3-1994055847656-1678695847103CMEAEYUfd79fef3d01d5e9aadc18ccd4d0c95073890; ci=20; _ga=GA1.1.397101976.1678695954; _ga_95GX0SH5GM=GS1.1.1678695953.1.0.1678695958.0.0.0; mtcdn=K; userTicket=oFCsIDbWMVgPoGuIOtmLyIceGDLeTsdKsJxTcZKs; _yoda_verify_resp=AJ%2FZONg0W7k05tIv8eL7iz4rH10JiiDLZ8OfyJDfkJiTCNurQq9K1NYSP2gkOH8c3DLXmhEZcZqX7xAsFOmj%2BEWjvxBUK%2F3VqE%2BKWZVpFKWaw6gz3uQbl9QQkMt%2FBMKVDesZPA7Sx0xu7%2Bc4KzZC6SzAoXa6c3QGsMPHKNHl2RGoFuzzlnEbrKUXcDqWiWIaM37HerrHsJWoxhiGltAPLrG9ftWXfj5fjJBSFvQAvPYq0H6pKafg7m1VDn4cRGsppf5E0Ou2GdjHI92C0h0perM91I%2FbKn7BwzJkVmPn6GJmb%2F0icYjjHGWt6NEia03rB7Zgse%2F8CdKVJoaxsA%2BmIccShXqihf65mb0N%2FDfA9uDKZ85QjeNTwM7BOTvStchP; _yoda_verify_rid=16b161df64435062; u=2654185033; n=vaU611418258; lt=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; mt_c_token=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; token=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; IJSESSIONID=node01g7w17ehbpc6n17bmml8i7xwi30735626; iuuid=EA94A2CA6CC7F75170785949EDBB804D251A1C6094D1F505772F701EFB35D117; isid=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; oops=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; logintype=normal; cityname=%E5%B9%BF%E5%B7%9E; _lxsdk=EA94A2CA6CC7F75170785949EDBB804D251A1C6094D1F505772F701EFB35D117; client-id=1fed5072-8319-43f2-a88d-072f7949c580; token2=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; _hc.v=8ccdbe4f-4e14-209a-0748-f81a4b45f65b.1678696143; unc=vaU611418258; lat=23.120109; lng=113.60865; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; firstTime=1678696167931; __mta=146117515.1678695942102.1678695942102.1678696167999.2; _lxsdk_s=186da11d302-caf-585-f21%7C%7C39',
          'Pragma': 'no-cache',
          'Referer': 'https://gz.meituan.com/meishi/rating/',
          'Sec-Fetch-Dest': 'document',
          'Sec-Fetch-Mode': 'navigate',
          'Sec-Fetch-Site': 'same-origin',
          'Sec-Fetch-User': '?1',
          'Upgrade-Insecure-Requests': '1',
          'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
          'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
          'sec-ch-ua-mobile': '?0',
          'sec-ch-ua-platform': '"Windows"'
        }
        proxies = get_ip_1()
        response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies)
        list_url = re.findall('{"poiId":(.*?),',response.text)
        for data in list_url:
          goods_url = 'https://gz.meituan.com/meishi/' + data + '/'
          redis_client.lpush('men_tuan',goods_url)
          print(f'存储成功data =={goods_url}')

def index():
    while True:
          # url = "https://gz.meituan.com/meishi/1907151454/"
          url = redis_client.lpop('men_tuan').decode("utf-8")
          payload = {}
          headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive',
            'Cookie': 'uuid=d3241f9aa4b04b23b78a.1678695846.1.0.0; _lxsdk_cuid=186da11d302c8-0a33418a36f0f-26031951-1fa400-186da11d302c8; WEBDFPID=70w500y31wvv59vu125w6u2vv1ux879w813wz88w30w979583190v6u3-1994055847656-1678695847103CMEAEYUfd79fef3d01d5e9aadc18ccd4d0c95073890; ci=20; _ga=GA1.1.397101976.1678695954; _ga_95GX0SH5GM=GS1.1.1678695953.1.0.1678695958.0.0.0; mtcdn=K; userTicket=oFCsIDbWMVgPoGuIOtmLyIceGDLeTsdKsJxTcZKs; _yoda_verify_resp=AJ%2FZONg0W7k05tIv8eL7iz4rH10JiiDLZ8OfyJDfkJiTCNurQq9K1NYSP2gkOH8c3DLXmhEZcZqX7xAsFOmj%2BEWjvxBUK%2F3VqE%2BKWZVpFKWaw6gz3uQbl9QQkMt%2FBMKVDesZPA7Sx0xu7%2Bc4KzZC6SzAoXa6c3QGsMPHKNHl2RGoFuzzlnEbrKUXcDqWiWIaM37HerrHsJWoxhiGltAPLrG9ftWXfj5fjJBSFvQAvPYq0H6pKafg7m1VDn4cRGsppf5E0Ou2GdjHI92C0h0perM91I%2FbKn7BwzJkVmPn6GJmb%2F0icYjjHGWt6NEia03rB7Zgse%2F8CdKVJoaxsA%2BmIccShXqihf65mb0N%2FDfA9uDKZ85QjeNTwM7BOTvStchP; _yoda_verify_rid=16b161df64435062; u=2654185033; n=vaU611418258; lt=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; mt_c_token=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; token=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; IJSESSIONID=node01g7w17ehbpc6n17bmml8i7xwi30735626; iuuid=EA94A2CA6CC7F75170785949EDBB804D251A1C6094D1F505772F701EFB35D117; isid=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; oops=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; logintype=normal; cityname=%E5%B9%BF%E5%B7%9E; _lxsdk=EA94A2CA6CC7F75170785949EDBB804D251A1C6094D1F505772F701EFB35D117; client-id=1fed5072-8319-43f2-a88d-072f7949c580; token2=AgFEINOmWHoJe2482squQ2EMdWOzjxD0jXMGxT2RKMvEJrsRmHn5-gphnRNfHBwxEvfl2u5qNSDmqgAAAAAGFwAAQ41oDwwNzEZfDahOeq7toSfW6sh6m6kt8O5PJI24QFA4b-IBaHRXMsBsR_PJCaE5; _hc.v=8ccdbe4f-4e14-209a-0748-f81a4b45f65b.1678696143; unc=vaU611418258; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; __mta=146117515.1678695942102.1678695942102.1678696167999.2; lat=30.310713; lng=120.327583; firstTime=1678697133157; _lxsdk_s=186da11d302-caf-585-f21%7C%7C61; client-id=1fed5072-8319-43f2-a88d-072f7949c580',
            'Pragma': 'no-cache',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Sec-Fetch-User': '?1',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Chromium";v="110", "Not A(Brand";v="24", "Google Chrome";v="110"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"'
          }
          proxies = get_ip_1()
          response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
          try:
            s = '{' + response.split('window._appState = {')[1].split(';</script>')[0]
          except:
            print(f'规则提取有误url==={url}')
            continue
          try:
            data = {}
            deals = []
            l = json.loads(s)
            shop = l['detailInfo']['name']
            address = l['detailInfo']['address']
            openTime = l['detailInfo']['openTime']
            phone = l['detailInfo']['phone']
            dealList = l['dealList']['deals']
            frontImgUrl = l['photos']['frontImgUrl']
            for deal in dealList:
              deal = deal['title']
              deals.append(deal)
            data['店铺名'] = shop
            data['地址'] = address
            data['营业时间'] = openTime
            data['联系电话'] = phone
            data['主营商品'] = deals
            data['iamge'] = frontImgUrl
            redis_client.hset(f"men_tun",url, json.dumps(data))
            print(f'存储成功data = {data}')
          except:
            print(f'出现问题url==={url}')
            continue

def save_goods_info():
    wb = load_workbook(f'./测试文件.xlsx')
    key_list = redis_client.hkeys(f"men_tun")
    sheet = wb['Sheet1']
    sheet.cell(row=1, column=2 ).value = "店铺名"
    sheet.cell(row=1, column=3 ).value = "地址"
    sheet.cell(row=1, column=4 ).value = "营业时间"
    sheet.cell(row=1, column=5 ).value = "联系电话"
    sheet.cell(row=1, column=6).value = "主营商品"
    sheet.cell(row=1, column=7).value = "iamge"
    i = 2
    for key in key_list:
        try:
            info = redis_client.hget(f"men_tun",key).decode("utf-8")
            data = json.loads(info)
            sheet.cell(row=i, column=2).value = data['店铺名']
            sheet.cell(row=i, column=3).value = data['地址']
            sheet.cell(row=i, column=4).value = data['营业时间']
            sheet.cell(row=i, column=5).value = data['联系电话']
            sheet.cell(row=i, column=6).value = str(data['主营商品'])
            sheet.cell(row=i, column=7).value = data['iamge']
            print(f"数据插入成功 item == {data}")
            i += 1
        except:
            print(f'出现错误，url=={data}')
            continue
    wb.save(f'./测试文件2.xlsx')

save_goods_info()