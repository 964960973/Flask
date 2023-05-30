import requests
import time
import json
import redis
import random
import threading
from openpyxl import load_workbook
redis_client = redis.Redis(host='127.0.0.1', port=6379)

def get_ip():
  proxy = ['http://5D1OsnsQ6JUZ:q5afKCYcs6ru@36.133.104.63:32080', 'http://wSszPfanZ5eW:641qnuJTJNSz@36.133.104.26:32080','http://MGv3HmBHv7Yb:AU4ZVU5DVkDx@36.133.103.174:32080','http://nC5D16KxZ5NA:4KkAOSHqEAFg@36.133.209.53:32080', 'http://YENEBBM7ZTSa:b60RKbb69w4P@36.138.63.10:32080','http://Jqg8E8FQEvP3:nXMmX7h5zKyT@36.133.103.186:32080']
  proxy = random.choice(proxy)
  return proxy

wb = load_workbook(f'./boss_test.xlsx')
def get_info_ex(CategoryName):
    url_list = redis_client.hkeys(f'tx_zp_ex{CategoryName}')
    i = 2
    sheet = wb[CategoryName]
    sheet.cell(row=1, column=1).value = "网址"
    sheet.cell(row=1, column=2).value = "岗位名称"
    sheet.cell(row=1, column=3).value = "岗位要求"
    sheet.cell(row=1, column=4).value = "入职工作"
    sheet.cell(row=1, column=5).value = "发布时间"
    sheet.cell(row=1, column=6).value = '城市'
    for key in url_list:
        try:
            data_info = redis_client.hget(f'tx_zp_ex{CategoryName}', key)
            info_str = data_info.decode("utf-8")
            info = json.loads(info_str)
            sheet.cell(row=i, column=1).value = key
            sheet.cell(row=i, column=2).value = info["RecruitPostName"]
            sheet.cell(row=i, column=3).value = info["Responsibility"]
            sheet.cell(row=i, column=2).value = info["Requirement"]
            sheet.cell(row=i, column=3).value = info["LastUpdateTime"]
            sheet.cell(row=i, column=4).value = info['LocationName']
            i += 1
            print(f"数据保存ex成功 url == {key}")
        except:
            print('出现错误')
            continue
    wb.save(f'腾讯招聘招聘信息.xlsx')

def get_info_two(CategoryName):
    while True:
        try:
            data = redis_client.lpop(f'tx_zp').decode("utf-8")
            zhu_ye = str(data).split('_')[0]
            id = str(data).split('_')[1]
            url = f"https://careers.tencent.com/tencentcareer/api/post/ByPostId?timestamp=1664180852579&postId={id}&language=zh-cn"
            payload = {}
            headers = {
                'authority': 'careers.tencent.com',
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'zh-CN,zh;q=0.9',
                'cache-control': 'no-cache',
                'cookie': 'sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24search_keyword_id%22%3A%2286ab49a20000122e00000006632d25e0%22%2C%22%24search_keyword_id_type%22%3A%22baidu_seo_keyword_id%22%2C%22%24search_keyword_id_hash%22%3A3609344412316440%2C%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTgzNjQzMDg4NzE1ODAtMGMyMzFkZDk1ZjJhN2Q4LTI2MDIxYzUxLTIwNzM2MDAtMTgzNjQzMDg4NzJiMjEifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%7D',
                'pragma': 'no-cache',
                'referer': 'https://careers.tencent.com/jobdesc.html',
                'sec-ch-ua': '"Google Chrome";v="105", "Not)A;Brand";v="8", "Chromium";v="105"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'
            }
            proxy = get_ip()
            proxies = {"http": proxy, "https": proxy}
            response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
            data_json = json.loads(response)
            item = {}
            item['RecruitPostName'] = data_json['Data']['RecruitPostName'] #岗位名称
            item['Responsibility'] = data_json['Data']['Responsibility']#岗位要求
            item['Requirement'] = data_json['Data']['Requirement']#入职工作
            item['LastUpdateTime'] = data_json['Data']['LastUpdateTime']#发布时间
            item['LocationName'] = data_json['Data']['LocationName']
            redis_client.hset(f"tx_zp_ex{CategoryName}", zhu_ye, json.dumps(item))
            print(f'上传数据库成功==={zhu_ye}')
        except:
            redis_client.rpush('tx_zp',data)
            print('当前链接全部获取完成')
            break

def get_timestamp():
    timestamp = int(time.time() * 1000)  # 获取13位时间戳
    return timestamp

def get_info(CategoryId):
    for page in range(1,10):
        time_time = get_timestamp()
        url = f"https://careers.tencent.com/tencentcareer/api/post/Query?timestamp={time_time}&countryId=&cityId=&bgIds=&productId=&categoryId=&parentCategoryId={CategoryId}&attrId=&keyword=&pageIndex={page}&pageSize=10&language=zh-cn&area=cn"
        payload={}
        headers = {
          'authority': 'careers.tencent.com',
          'accept': 'application/json, text/plain, */*',
          'accept-language': 'zh-CN,zh;q=0.9',
          'cache-control': 'no-cache',
          'cookie': 'sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24search_keyword_id%22%3A%2286ab49a20000122e00000006632d25e0%22%2C%22%24search_keyword_id_type%22%3A%22baidu_seo_keyword_id%22%2C%22%24search_keyword_id_hash%22%3A3609344412316440%2C%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTgzNjQzMDg4NzE1ODAtMGMyMzFkZDk1ZjJhN2Q4LTI2MDIxYzUxLTIwNzM2MDAtMTgzNjQzMDg4NzJiMjEifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%7D',
          'pragma': 'no-cache',
          'referer': 'https://careers.tencent.com/search.html?pcid=40001',
          'sec-ch-ua': '"Google Chrome";v="105", "Not)A;Brand";v="8", "Chromium";v="105"',
          'sec-ch-ua-mobile': '?0',
          'sec-ch-ua-platform': '"Windows"',
          'sec-fetch-dest': 'empty',
          'sec-fetch-mode': 'cors',
          'sec-fetch-site': 'same-origin',
          'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'
        }
        proxy = get_ip()
        proxies = {"http": proxy, "https": proxy}
        response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
        try:
            json_data = json.loads(response)
            if json_data['Data']['Count'] == 0:
                print('获取完成')
                break
            else:
                Data = json_data['Data']['Posts']
                for item in Data:
                    PostURL = item['PostURL']
                    RecruitPostId = item['RecruitPostId']
                    redis_client.lpush(f'tx_zp',PostURL+'_'+str(RecruitPostId))
                    print(PostURL+'获取成功')
        except:
            print(f'上传失败url==={url}')

def get_index():
    time_time = get_timestamp()
    url = f"https://careers.tencent.com/tencentcareer/api/post/ByCategories?timestamp={time_time}&language=zh-cn"
    payload = {}
    headers = {
        'authority': 'careers.tencent.com',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cache-control': 'no-cache',
        'cookie': 'sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24search_keyword_id%22%3A%2286ab49a20000122e00000006632d25e0%22%2C%22%24search_keyword_id_type%22%3A%22baidu_seo_keyword_id%22%2C%22%24search_keyword_id_hash%22%3A3609344412316440%2C%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTgzNjQzMDg4NzE1ODAtMGMyMzFkZDk1ZjJhN2Q4LTI2MDIxYzUxLTIwNzM2MDAtMTgzNjQzMDg4NzJiMjEifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218364308871580-0c231dd95f2a7d8-26021c51-2073600-18364308872b21%22%7D',
        'pragma': 'no-cache',
        'referer': 'https://careers.tencent.com/jobopportunity.html',
        'sec-ch-ua': '"Google Chrome";v="105", "Not)A;Brand";v="8", "Chromium";v="105"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'
    }
    proxy = get_ip()
    proxies = {"http": proxy, "https": proxy}
    response = requests.request("GET", url, headers=headers, data=payload,proxies=proxies).text
    json_data = json.loads(response)
    Data = json_data['Data']
    for data in Data:
        CategoryId = data['CategoryId']
        time.sleep(2)
        print('正在进行提取操作，请耐心等待')
        get_info(CategoryId)


if __name__ == '__main__':
    # get_index() #先运行这个
    # main() #然后这个
    list = ['技术类','产品类','内容类','设计类','销售、服务与支持类','人力资源类','营销与公关类','战略与投资类','财务类','法律与公共策略类','行政支持类']
    for i in list:
        # get_info_two(i)
        get_info_ex(i)
        # pass
    print('已经全部获取完成!!!!!!!!!!!!!')