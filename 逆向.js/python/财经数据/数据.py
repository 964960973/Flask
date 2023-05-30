#!/usr/bin/env python
# coding=utf-8
from openpyxl import load_workbook
import re
import requests
import time

times = time.localtime().tm_year
ti = time.localtime().tm_mon
t = time.localtime().tm_mday
year = str(times)+'_'+str(ti)+'_'+str(t)



def get_finance(country,col_num):
        # url = f"https://vip.stock.finance.sina.com.cn/forex/api/jsonp.php/var%20_{country}{year}=/NewForexService.getDayKLine?symbol={country}&_={year}"
        url = f"https://vip.stock.finance.sina.com.cn/forex/api/jsonp.php/var%20_{country}{year}=/NewForexService.getDayKLine?symbol={country}&_={year}"
        payload = {}
        headers = {
            'authority': 'vip.stock.finance.sina.com.cn',
            'accept': '*/*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'no-cache',
            'cookie': 'UOR=,finance.sina.com.cn,; U_TRS1=00000065.d0cc1287.63436f2b.4c768f84; U_TRS2=00000065.d0d51287.63436f2b.4b5cb3aa; SINAGLOBAL=173.242.113.101_1665363774.44387; Apache=173.242.113.101_1665363774.44389; MONEY-FINANCE-SINA-COM-CN-WEB5=; ULV=1665363839753:2:2:2:173.242.113.101_1665363774.44389:1665363751779; MONEY-FINANCE-SINA-COM-CN-WEB5=',
            'pragma': 'no-cache',
            'referer': 'https://finance.sina.com.cn/money/forex/hq/USDCNY.shtml',
            'sec-ch-ua': '"Google Chrome";v="105", "Not)A;Brand";v="8", "Chromium";v="105"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'script',
            'sec-fetch-mode': 'no-cors',
            'sec-fetch-site': 'same-site',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'
        }
        response = requests.request("GET", url, headers=headers).text
        # print(response)
        l =response.split('=(')[1]
        k = l.split('");')[0] + '}}}'
        # '2022-10-03'
        try:
            s = re.findall(r'2023-04-01(.*?)}}}',k)
            s = '|2023-04-01' + s[0]
        except:
            if '2023-04-02' not in response:
                s = re.findall(r'2023-04-03(.*?)}}}', k)
                s = '|2023-04-02' + s[0]
            else:
                s = re.findall(r'2023-04-02(.*?)}}}', k)
                s = '|2023-04-02' + s[0]
        data = s.split("|")
        # print(data)
        data.pop(0)
        a = 3
        day = 1
        col = int(col_num) * 4
        wb = load_workbook(f'./测试文件.xlsx')  # 加载excel  要先存在 作为模板
        sheet = wb.active
        sheet.cell(row=1, column=1 + col).value = country
        sheet.cell(row=2, column=1 + col).value = "data"
        sheet.cell(row=2, column=2 + col).value = "max"
        sheet.cell(row=2, column=3 + col).value = "min"
        # 取默认的sheet  不用管
        for i in data:
            while True:
                # print(i)
                info_list = i.split(",")
                timestap = info_list.pop(0)
                if day < 10:
                    day_str = "0"+str(day)
                else:
                    day_str = str(day)
                if day > 31:
                    break
                if f"2023-04-{day_str}" != timestap:
                    a += 1
                    day += 1
                    continue
                info_list.pop(-1)
                max_num = max(info_list)
                min_num = min(info_list)
                sheet.cell(row=a, column=int(1)+col).value = str(timestap)
                sheet.cell(row=a, column=int(2)+col).value = max_num
                sheet.cell(row=a, column=int(3)+col).value = min_num
                a += 1
                day += 1
                print("数据插入成功")
                wb.save('./测试文件.xlsx')
                # wb.save(f'test1.xlsx')
                break
col_num = 0
list = ["fx_smxncny", "fx_susdcny", "fx_scadcny"]
# list = ["fx_susdcny"]
for country in list:
    get_finance(country,str(col_num))
    col_num += 1
